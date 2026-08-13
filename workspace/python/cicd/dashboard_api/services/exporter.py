import io
import csv
import json
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors


def export_csv(data: dict) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Project", "Pipeline ID", "Status", "Ref", "Duration (s)", "Created At", "URL"])
    for p in data.get("pipelines", []):
        writer.writerow([
            p.get("project_name", ""),
            p.get("id", ""),
            p.get("status", ""),
            p.get("ref", ""),
            p.get("duration", ""),
            p.get("created_at", ""),
            p.get("web_url", ""),
        ])
    return output.getvalue().encode()


def export_json(data: dict) -> bytes:
    envelope = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "source": "cicd-dashboard",
        "data": data,
    }
    return json.dumps(envelope, indent=2, default=str).encode()


def _style_sheet(ws):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F4F6")
    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
            elif i % 2 == 0:
                cell.fill = alt_fill
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def export_excel(data: dict) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pipelines = data.get("pipelines", [])
        if pipelines:
            df = pd.DataFrame([{
                "Project": p.get("project_name", ""),
                "ID": p.get("id", ""),
                "Status": p.get("status", ""),
                "Ref": p.get("ref", ""),
                "Duration (s)": p.get("duration", ""),
                "Created At": p.get("created_at", ""),
                "URL": p.get("web_url", ""),
            } for p in pipelines])
            df.to_excel(writer, sheet_name="Pipelines", index=False)
            _style_sheet(writer.sheets["Pipelines"])

        mrs = data.get("merge_requests", [])
        if mrs:
            df = pd.DataFrame([{
                "Project": m.get("project_name", ""),
                "ID": m.get("id", ""),
                "Title": m.get("title", ""),
                "Author": m.get("author_name", m.get("author", "")),
                "State": m.get("state", ""),
                "Created At": m.get("created_at", ""),
            } for m in mrs])
            df.to_excel(writer, sheet_name="Merge Requests", index=False)
            _style_sheet(writer.sheets["Merge Requests"])

    output.seek(0)
    return output


def export_pdf(data: dict) -> io.BytesIO:
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("CI/CD Dashboard Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    pipelines = data.get("pipelines", [])[:50]
    if pipelines:
        elements.append(Paragraph("Pipelines", styles["Heading2"]))
        rows = [["Project", "ID", "Status", "Ref", "Duration", "Created"]]
        for p in pipelines:
            dur = p.get("duration")
            rows.append([
                str(p.get("project_name", ""))[:22],
                str(p.get("id", "")),
                str(p.get("status", "")),
                str(p.get("ref", ""))[:20],
                f"{int(dur)}s" if dur else "-",
                str(p.get("created_at", ""))[:10],
            ])
        t = Table(rows)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

    doc.build(elements)
    output.seek(0)
    return output
