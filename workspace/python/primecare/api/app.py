import os
import io
import csv
import json
from datetime import date, time, datetime, timedelta
from decimal import Decimal
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import psycopg2
import psycopg2.extras

app = Flask(__name__)
CORS(app)

# Set DB_PASSWORD before starting:
#   Windows (PowerShell): $env:DB_PASSWORD="yourpassword"
#   macOS/Linux:          export DB_PASSWORD="yourpassword"
DB = dict(
    host=os.getenv('DB_HOST', 'localhost'),
    port=int(os.getenv('DB_PORT', '5432')),
    dbname=os.getenv('DB_NAME', 'clinic'),
    user=os.getenv('DB_USER', 'postgres'),
    password=os.getenv('DB_PASSWORD', ''),
)


def get_conn():
    return psycopg2.connect(**DB)


def query(sql, params=None):
    c = get_conn()
    try:
        cur = c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params)
        c.commit()
        if cur.description:
            result = [_clean(dict(row)) for row in cur.fetchall()]
        else:
            result = {'rows': cur.rowcount}
        cur.close()
        return result
    except Exception:
        c.rollback()
        raise
    finally:
        c.close()


def _clean(row):
    for k, v in row.items():
        if isinstance(v, (datetime, date)):
            row[k] = v.isoformat()
        elif isinstance(v, time):
            row[k] = v.strftime('%H:%M:%S')
        elif isinstance(v, Decimal):
            row[k] = float(v)
    return row


def ok(data, code=200):
    return jsonify(data), code


def fail(msg, code=400):
    return jsonify({'error': msg}), code


@app.errorhandler(Exception)
def handle_exception(e):
    return fail(str(e), 500)


# ── SPECIALITIES ──────────────────────────────────────────────────────────────

@app.route('/specialities', methods=['GET'])
def list_specialities():
    return ok(query('SELECT * FROM specialities ORDER BY name'))


@app.route('/specialities/<int:id>', methods=['GET'])
def get_speciality(id):
    rows = query('SELECT * FROM specialities WHERE id=%s', (id,))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/specialities', methods=['POST'])
def create_speciality():
    d = request.json
    rows = query('INSERT INTO specialities (name) VALUES (%s) RETURNING *', (d['name'],))
    return ok(rows[0], 201)


@app.route('/specialities/<int:id>', methods=['PUT'])
def update_speciality(id):
    d = request.json
    rows = query('UPDATE specialities SET name=%s WHERE id=%s RETURNING *', (d['name'], id))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/specialities/<int:id>', methods=['DELETE'])
def delete_speciality(id):
    query('DELETE FROM specialities WHERE id=%s', (id,))
    return ok({'deleted': id})


# ── DOCTORS ───────────────────────────────────────────────────────────────────

DOCTOR_SELECT = '''
    SELECT d.*, s.name AS speciality_name
    FROM doctors d
    JOIN specialities s ON s.id = d.speciality_id
'''


@app.route('/doctors', methods=['GET'])
def list_doctors():
    return ok(query(DOCTOR_SELECT + 'ORDER BY d.last_name'))


@app.route('/doctors/<int:id>', methods=['GET'])
def get_doctor(id):
    rows = query(DOCTOR_SELECT + 'WHERE d.id=%s', (id,))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/doctors', methods=['POST'])
def create_doctor():
    d = request.json
    rows = query('''
        INSERT INTO doctors
            (first_name, last_name, email, phone, registration_number,
             speciality_id, bio, consultation_fee, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *
    ''', (d['first_name'], d['last_name'], d['email'], d['phone'],
          d['registration_number'], d['speciality_id'], d.get('bio'),
          d.get('consultation_fee', 500), d.get('is_active', True)))
    return ok(rows[0], 201)


@app.route('/doctors/<int:id>', methods=['PUT'])
def update_doctor(id):
    d = request.json
    rows = query('''
        UPDATE doctors SET
            first_name=%s, last_name=%s, email=%s, phone=%s,
            registration_number=%s, speciality_id=%s, bio=%s,
            consultation_fee=%s, is_active=%s
        WHERE id=%s RETURNING *
    ''', (d['first_name'], d['last_name'], d['email'], d['phone'],
          d['registration_number'], d['speciality_id'], d.get('bio'),
          d.get('consultation_fee', 500), d.get('is_active', True), id))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/doctors/<int:id>', methods=['DELETE'])
def delete_doctor(id):
    query('DELETE FROM doctors WHERE id=%s', (id,))
    return ok({'deleted': id})


# ── DOCTOR SCHEDULES ──────────────────────────────────────────────────────────

@app.route('/schedules', methods=['GET'])
def list_schedules():
    return ok(query('SELECT * FROM doctor_schedules ORDER BY doctor_id, day_of_week, start_time'))


@app.route('/schedules/<int:id>', methods=['GET'])
def get_schedule(id):
    rows = query('SELECT * FROM doctor_schedules WHERE id=%s', (id,))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/schedules', methods=['POST'])
def create_schedule():
    d = request.json
    rows = query('''
        INSERT INTO doctor_schedules
            (doctor_id, day_of_week, start_time, end_time, slot_duration_minutes, is_active)
        VALUES (%s, %s, %s, %s, %s, %s) RETURNING *
    ''', (d['doctor_id'], d['day_of_week'], d['start_time'], d['end_time'],
          d.get('slot_duration_minutes', 30), d.get('is_active', True)))
    return ok(rows[0], 201)


@app.route('/schedules/<int:id>', methods=['PUT'])
def update_schedule(id):
    d = request.json
    rows = query('''
        UPDATE doctor_schedules SET
            doctor_id=%s, day_of_week=%s, start_time=%s, end_time=%s,
            slot_duration_minutes=%s, is_active=%s
        WHERE id=%s RETURNING *
    ''', (d['doctor_id'], d['day_of_week'], d['start_time'], d['end_time'],
          d.get('slot_duration_minutes', 30), d.get('is_active', True), id))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/schedules/<int:id>', methods=['DELETE'])
def delete_schedule(id):
    query('DELETE FROM doctor_schedules WHERE id=%s', (id,))
    return ok({'deleted': id})


# ── PATIENTS ──────────────────────────────────────────────────────────────────

@app.route('/patients', methods=['GET'])
def list_patients():
    return ok(query('SELECT * FROM patients ORDER BY last_name'))


@app.route('/patients/<int:id>', methods=['GET'])
def get_patient(id):
    rows = query('SELECT * FROM patients WHERE id=%s', (id,))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/patients', methods=['POST'])
def create_patient():
    d = request.json
    rows = query('''
        INSERT INTO patients (first_name, last_name, email, phone, date_of_birth, gender)
        VALUES (%s, %s, %s, %s, %s, %s) RETURNING *
    ''', (d['first_name'], d['last_name'], d.get('email'), d['phone'],
          d.get('date_of_birth'), d.get('gender')))
    return ok(rows[0], 201)


@app.route('/patients/<int:id>', methods=['PUT'])
def update_patient(id):
    d = request.json
    rows = query('''
        UPDATE patients SET
            first_name=%s, last_name=%s, email=%s, phone=%s,
            date_of_birth=%s, gender=%s
        WHERE id=%s RETURNING *
    ''', (d['first_name'], d['last_name'], d.get('email'), d['phone'],
          d.get('date_of_birth'), d.get('gender'), id))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/patients/<int:id>', methods=['DELETE'])
def delete_patient(id):
    query('DELETE FROM patients WHERE id=%s', (id,))
    return ok({'deleted': id})


# ── APPOINTMENTS ──────────────────────────────────────────────────────────────

APPT_SELECT = '''
    SELECT a.*,
           d.first_name || ' ' || d.last_name AS doctor_name,
           p.first_name || ' ' || p.last_name AS patient_name,
           s.name AS speciality_name
    FROM appointments a
    JOIN doctors  d ON d.id = a.doctor_id
    JOIN patients p ON p.id = a.patient_id
    JOIN specialities s ON s.id = d.speciality_id
'''


@app.route('/appointments', methods=['GET'])
def list_appointments():
    return ok(query(APPT_SELECT + 'ORDER BY a.appointment_date, a.start_time'))


@app.route('/appointments/<int:id>', methods=['GET'])
def get_appointment(id):
    rows = query(APPT_SELECT + 'WHERE a.id=%s', (id,))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/appointments', methods=['POST'])
def create_appointment():
    d = request.json
    rows = query('''
        INSERT INTO appointments
            (doctor_id, patient_id, appointment_date, start_time, end_time, status)
        VALUES (%s, %s, %s, %s, %s, %s) RETURNING *
    ''', (d['doctor_id'], d['patient_id'], d['appointment_date'],
          d['start_time'], d['end_time'], d.get('status', 'confirmed')))
    return ok(rows[0], 201)


@app.route('/appointments/<int:id>', methods=['PUT'])
def update_appointment(id):
    d = request.json
    rows = query('''
        UPDATE appointments SET
            doctor_id=%s, patient_id=%s, appointment_date=%s,
            start_time=%s, end_time=%s, status=%s
        WHERE id=%s RETURNING *
    ''', (d['doctor_id'], d['patient_id'], d['appointment_date'],
          d['start_time'], d['end_time'], d.get('status', 'confirmed'), id))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/appointments/<int:id>', methods=['DELETE'])
def delete_appointment(id):
    query('DELETE FROM appointments WHERE id=%s', (id,))
    return ok({'deleted': id})


# ── CASE HISTORY ──────────────────────────────────────────────────────────────

@app.route('/case-history', methods=['GET'])
def list_case_history():
    return ok(query('SELECT * FROM case_history ORDER BY created_at DESC'))


@app.route('/case-history/<int:id>', methods=['GET'])
def get_case_history(id):
    rows = query('SELECT * FROM case_history WHERE id=%s', (id,))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/case-history', methods=['POST'])
def create_case_history():
    d = request.json
    rows = query('''
        INSERT INTO case_history
            (appointment_id, symptoms, diagnosis, prescription,
             follow_up_needed, follow_up_date, follow_up_appointment_id, notes)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING *
    ''', (d['appointment_id'], d.get('symptoms'), d.get('diagnosis'),
          d.get('prescription'), d.get('follow_up_needed', False),
          d.get('follow_up_date'), d.get('follow_up_appointment_id'), d.get('notes')))
    return ok(rows[0], 201)


@app.route('/case-history/<int:id>', methods=['PUT'])
def update_case_history(id):
    d = request.json
    rows = query('''
        UPDATE case_history SET
            symptoms=%s, diagnosis=%s, prescription=%s,
            follow_up_needed=%s, follow_up_date=%s,
            follow_up_appointment_id=%s, notes=%s
        WHERE id=%s RETURNING *
    ''', (d.get('symptoms'), d.get('diagnosis'), d.get('prescription'),
          d.get('follow_up_needed', False), d.get('follow_up_date'),
          d.get('follow_up_appointment_id'), d.get('notes'), id))
    return ok(rows[0]) if rows else fail('Not found', 404)


@app.route('/case-history/<int:id>', methods=['DELETE'])
def delete_case_history(id):
    query('DELETE FROM case_history WHERE id=%s', (id,))
    return ok({'deleted': id})


# ── SPECIAL ENDPOINTS ─────────────────────────────────────────────────────────

@app.route('/doctors/by-speciality/<int:speciality_id>', methods=['GET'])
def doctors_by_speciality(speciality_id):
    return ok(query(
        DOCTOR_SELECT + 'WHERE d.speciality_id=%s ORDER BY d.last_name',
        (speciality_id,)
    ))


@app.route('/schedules/by-doctor/<int:doctor_id>', methods=['GET'])
def schedules_by_doctor(doctor_id):
    return ok(query('''
        SELECT ds.*, d.first_name || ' ' || d.last_name AS doctor_name
        FROM doctor_schedules ds
        JOIN doctors d ON d.id = ds.doctor_id
        WHERE ds.doctor_id=%s
        ORDER BY ds.day_of_week, ds.start_time
    ''', (doctor_id,)))


@app.route('/schedules/by-speciality/<int:speciality_id>', methods=['GET'])
def schedules_by_speciality(speciality_id):
    return ok(query('''
        SELECT ds.*, d.first_name || ' ' || d.last_name AS doctor_name, s.name AS speciality_name
        FROM doctor_schedules ds
        JOIN doctors d ON d.id = ds.doctor_id
        JOIN specialities s ON s.id = d.speciality_id
        WHERE d.speciality_id=%s
        ORDER BY ds.day_of_week, ds.start_time
    ''', (speciality_id,)))


@app.route('/doctors/by-slot', methods=['GET'])
def doctors_by_slot():
    day   = request.args.get('day_of_week')
    start = request.args.get('start_time')
    if not day or not start:
        return fail('day_of_week and start_time are required')
    return ok(query('''
        SELECT DISTINCT d.*, s.name AS speciality_name
        FROM doctors d
        JOIN specialities s ON s.id = d.speciality_id
        JOIN doctor_schedules ds ON ds.doctor_id = d.id
        WHERE ds.day_of_week = %s
          AND ds.start_time <= %s::time
          AND ds.end_time   >  %s::time
          AND ds.is_active = TRUE
          AND d.is_active  = TRUE
        ORDER BY d.last_name
    ''', (day, start, start)))


@app.route('/appointments/by-doctor/<int:doctor_id>', methods=['GET'])
def appointments_by_doctor(doctor_id):
    return ok(query(
        APPT_SELECT + 'WHERE a.doctor_id=%s ORDER BY a.appointment_date, a.start_time',
        (doctor_id,)
    ))


@app.route('/patients/by-doctor/<int:doctor_id>', methods=['GET'])
def patients_by_doctor(doctor_id):
    return ok(query('''
        SELECT DISTINCT p.*
        FROM patients p
        JOIN appointments a ON a.patient_id = p.id
        WHERE a.doctor_id=%s
        ORDER BY p.last_name
    ''', (doctor_id,)))


@app.route('/appointments/book', methods=['POST'])
def book_appointment():
    d          = request.json
    patient_id = d.get('patient_id')
    doctor_id  = d.get('doctor_id')
    appt_date  = d.get('appointment_date')
    start_time = d.get('start_time')
    end_time   = d.get('end_time')

    if not all([patient_id, doctor_id, appt_date, start_time, end_time]):
        return fail('patient_id, doctor_id, appointment_date, start_time and end_time are required')

    if not query('SELECT id FROM patients WHERE id=%s', (patient_id,)):
        return fail('Patient not found', 404)

    if not query('SELECT id FROM doctors WHERE id=%s AND is_active=TRUE', (doctor_id,)):
        return fail('Doctor not found or inactive', 404)

    if not query('''
        SELECT id FROM doctor_schedules
        WHERE doctor_id  = %s
          AND day_of_week = EXTRACT(DOW FROM %s::date)::int
          AND start_time <= %s::time
          AND end_time   >= %s::time
          AND is_active  = TRUE
    ''', (doctor_id, appt_date, start_time, end_time)):
        return fail('Doctor has no schedule covering this slot', 422)

    if query('SELECT id FROM doctor_leaves WHERE doctor_id=%s AND leave_date=%s',
             (doctor_id, appt_date)):
        return fail('Doctor is on leave on this date', 422)

    if query('''
        SELECT id FROM appointments
        WHERE doctor_id=%s AND appointment_date=%s AND start_time=%s AND status != 'cancelled'
    ''', (doctor_id, appt_date, start_time)):
        return fail('Slot already booked', 409)

    rows = query('''
        INSERT INTO appointments
            (doctor_id, patient_id, appointment_date, start_time, end_time, status)
        VALUES (%s, %s, %s, %s, %s, 'confirmed') RETURNING *
    ''', (doctor_id, patient_id, appt_date, start_time, end_time))
    return ok(rows[0], 201)


# ── DOCTOR LEAVES ─────────────────────────────────────────────────────────────

@app.route('/leaves', methods=['GET'])
def list_leaves():
    return ok(query('''
        SELECT dl.*, d.first_name || ' ' || d.last_name AS doctor_name
        FROM doctor_leaves dl
        JOIN doctors d ON d.id = dl.doctor_id
        ORDER BY dl.leave_date DESC, d.last_name
    '''))


@app.route('/leaves/by-doctor/<int:doctor_id>', methods=['GET'])
def leaves_by_doctor(doctor_id):
    return ok(query(
        'SELECT * FROM doctor_leaves WHERE doctor_id=%s ORDER BY leave_date DESC',
        (doctor_id,)
    ))


@app.route('/leaves', methods=['POST'])
def create_leave():
    d          = request.json
    doctor_id  = d.get('doctor_id')
    leave_date = d.get('leave_date')
    if not doctor_id or not leave_date:
        return fail('doctor_id and leave_date are required')
    if not query('SELECT id FROM doctors WHERE id=%s', (doctor_id,)):
        return fail('Doctor not found', 404)
    try:
        rows = query('''
            INSERT INTO doctor_leaves (doctor_id, leave_date, reason)
            VALUES (%s, %s, %s) RETURNING *
        ''', (doctor_id, leave_date, d.get('reason')))
        return ok(rows[0], 201)
    except Exception as e:
        if 'unique' in str(e).lower():
            return fail('Leave already marked for this doctor on this date', 409)
        raise


@app.route('/leaves/<int:id>', methods=['DELETE'])
def delete_leave(id):
    query('DELETE FROM doctor_leaves WHERE id=%s', (id,))
    return ok({'deleted': id})


# ══════════════════════════════════════════════════════════════════════════════
# REPORTING MODULE
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/reports/summary', methods=['GET'])
def get_report_summary():
    """
    Get overall summary statistics for management dashboard.
    """
    try:
        # Total counts
        total_doctors = query('SELECT COUNT(*) as count FROM doctors WHERE is_active=TRUE')[0]['count']
        total_patients = query('SELECT COUNT(*) as count FROM patients')[0]['count']
        total_specialities = query('SELECT COUNT(*) as count FROM specialities')[0]['count']

        # Appointment statistics
        total_appointments = query('SELECT COUNT(*) as count FROM appointments')[0]['count']
        confirmed = query("SELECT COUNT(*) as count FROM appointments WHERE status='confirmed'")[0]['count']
        completed = query("SELECT COUNT(*) as count FROM appointments WHERE status='completed'")[0]['count']
        cancelled = query("SELECT COUNT(*) as count FROM appointments WHERE status='cancelled'")[0]['count']

        # Revenue statistics
        revenue_result = query('''
            SELECT COALESCE(SUM(d.consultation_fee), 0) as total_revenue
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            WHERE a.status = 'completed'
        ''')
        total_revenue = float(revenue_result[0]['total_revenue'])

        # This month's statistics
        month_appointments = query('''
            SELECT COUNT(*) as count FROM appointments
            WHERE EXTRACT(MONTH FROM appointment_date) = EXTRACT(MONTH FROM CURRENT_DATE)
              AND EXTRACT(YEAR FROM appointment_date) = EXTRACT(YEAR FROM CURRENT_DATE)
        ''')[0]['count']

        month_revenue = query('''
            SELECT COALESCE(SUM(d.consultation_fee), 0) as revenue
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            WHERE a.status = 'completed'
              AND EXTRACT(MONTH FROM a.appointment_date) = EXTRACT(MONTH FROM CURRENT_DATE)
              AND EXTRACT(YEAR FROM a.appointment_date) = EXTRACT(YEAR FROM CURRENT_DATE)
        ''')
        this_month_revenue = float(month_revenue[0]['revenue'])

        # Top specialities by appointments
        top_specialities = query('''
            SELECT s.name, COUNT(a.id) as appointment_count
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            JOIN specialities s ON s.id = d.speciality_id
            GROUP BY s.name
            ORDER BY appointment_count DESC
            LIMIT 5
        ''')

        # Recent appointments
        recent = query('''
            SELECT a.*, 
                   d.first_name || ' ' || d.last_name AS doctor_name,
                   p.first_name || ' ' || p.last_name AS patient_name,
                   s.name AS speciality_name
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            JOIN patients p ON p.id = a.patient_id
            JOIN specialities s ON s.id = d.speciality_id
            ORDER BY a.appointment_date DESC, a.start_time DESC
            LIMIT 10
        ''')

        return ok({
            'overview': {
                'total_doctors': total_doctors,
                'total_patients': total_patients,
                'total_specialities': total_specialities,
                'total_appointments': total_appointments,
                'total_revenue': total_revenue
            },
            'appointments': {
                'total': total_appointments,
                'confirmed': confirmed,
                'completed': completed,
                'cancelled': cancelled,
                'this_month': month_appointments
            },
            'revenue': {
                'total': total_revenue,
                'this_month': this_month_revenue
            },
            'top_specialities': top_specialities,
            'recent_appointments': recent
        })
    except Exception as e:
        return fail(str(e), 500)


@app.route('/reports/appointments', methods=['GET'])
def get_appointments_report():
    """
    Get detailed appointments report with optional filters.
    Query params: start_date, end_date, doctor_id, patient_id, status, speciality_id
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        doctor_id = request.args.get('doctor_id')
        patient_id = request.args.get('patient_id')
        status = request.args.get('status')
        speciality_id = request.args.get('speciality_id')

        sql = '''
            SELECT a.*,
                   d.first_name || ' ' || d.last_name AS doctor_name,
                   p.first_name || ' ' || p.last_name AS patient_name,
                   s.name AS speciality_name,
                   d.consultation_fee
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            JOIN patients p ON p.id = a.patient_id
            JOIN specialities s ON s.id = d.speciality_id
            WHERE 1=1
        '''
        params = []

        if start_date:
            params.append(start_date)
            sql += f' AND a.appointment_date >= ${len(params)}'
        if end_date:
            params.append(end_date)
            sql += f' AND a.appointment_date <= ${len(params)}'
        if doctor_id:
            params.append(int(doctor_id))
            sql += f' AND a.doctor_id = ${len(params)}'
        if patient_id:
            params.append(int(patient_id))
            sql += f' AND a.patient_id = ${len(params)}'
        if status:
            params.append(status)
            sql += f' AND a.status = ${len(params)}'
        if speciality_id:
            params.append(int(speciality_id))
            sql += f' AND d.speciality_id = ${len(params)}'

        sql += ' ORDER BY a.appointment_date DESC, a.start_time DESC'

        # Replace $1, $2 with %s for psycopg2
        for i in range(len(params), 0, -1):
            sql = sql.replace(f'${i}', '%s')

        appointments = query(sql, tuple(params) if params else None)

        # Calculate summary statistics
        total_revenue = sum(float(a['consultation_fee']) for a in appointments if a['status'] == 'completed')
        summary = {
            'total_appointments': len(appointments),
            'confirmed': sum(1 for a in appointments if a['status'] == 'confirmed'),
            'completed': sum(1 for a in appointments if a['status'] == 'completed'),
            'cancelled': sum(1 for a in appointments if a['status'] == 'cancelled'),
            'total_revenue': total_revenue
        }

        return ok({
            'summary': summary,
            'appointments': appointments
        })
    except Exception as e:
        return fail(str(e), 500)


@app.route('/reports/doctors', methods=['GET'])
def get_doctors_report():
    """
    Get doctor performance report.
    """
    try:
        doctors_stats = query('''
            SELECT 
                d.id,
                d.first_name || ' ' || d.last_name AS doctor_name,
                s.name AS speciality_name,
                d.consultation_fee,
                COUNT(CASE WHEN a.status = 'completed' THEN 1 END) as completed_appointments,
                COUNT(CASE WHEN a.status = 'confirmed' THEN 1 END) as upcoming_appointments,
                COUNT(CASE WHEN a.status = 'cancelled' THEN 1 END) as cancelled_appointments,
                COUNT(a.id) as total_appointments,
                COALESCE(SUM(CASE WHEN a.status = 'completed' THEN d.consultation_fee ELSE 0 END), 0) as total_revenue
            FROM doctors d
            JOIN specialities s ON s.id = d.speciality_id
            LEFT JOIN appointments a ON a.doctor_id = d.id
            WHERE d.is_active = TRUE
            GROUP BY d.id, d.first_name, d.last_name, s.name, d.consultation_fee
            ORDER BY total_revenue DESC
        ''')

        # Calculate average statistics
        total_docs = len(doctors_stats)
        if total_docs > 0:
            avg_appointments = sum(d['total_appointments'] for d in doctors_stats) / total_docs
            avg_revenue = sum(float(d['total_revenue']) for d in doctors_stats) / total_docs
        else:
            avg_appointments = 0
            avg_revenue = 0

        return ok({
            'summary': {
                'total_doctors': total_docs,
                'average_appointments_per_doctor': round(avg_appointments, 2),
                'average_revenue_per_doctor': round(avg_revenue, 2)
            },
            'doctors': doctors_stats
        })
    except Exception as e:
        return fail(str(e), 500)


@app.route('/reports/specialities', methods=['GET'])
def get_specialities_report():
    """
    Get speciality-wise report.
    """
    try:
        speciality_stats = query('''
            SELECT 
                s.id,
                s.name AS speciality_name,
                COUNT(DISTINCT d.id) as doctor_count,
                COUNT(a.id) as total_appointments,
                COUNT(CASE WHEN a.status = 'completed' THEN 1 END) as completed_appointments,
                COUNT(CASE WHEN a.status = 'confirmed' THEN 1 END) as upcoming_appointments,
                COALESCE(SUM(CASE WHEN a.status = 'completed' THEN d.consultation_fee ELSE 0 END), 0) as total_revenue,
                COALESCE(AVG(CASE WHEN a.status = 'completed' THEN d.consultation_fee END), 0) as avg_consultation_fee
            FROM specialities s
            LEFT JOIN doctors d ON d.speciality_id = s.id AND d.is_active = TRUE
            LEFT JOIN appointments a ON a.doctor_id = d.id
            GROUP BY s.id, s.name
            ORDER BY total_revenue DESC
        ''')

        return ok({
            'specialities': speciality_stats
        })
    except Exception as e:
        return fail(str(e), 500)


@app.route('/reports/patients', methods=['GET'])
def get_patients_report():
    """
    Get patient statistics report.
    """
    try:
        patient_stats = query('''
            SELECT 
                p.id,
                p.first_name || ' ' || p.last_name AS patient_name,
                p.gender,
                p.date_of_birth,
                EXTRACT(YEAR FROM AGE(p.date_of_birth)) as age,
                COUNT(a.id) as total_appointments,
                COUNT(CASE WHEN a.status = 'completed' THEN 1 END) as completed_appointments,
                COUNT(CASE WHEN a.status = 'confirmed' THEN 1 END) as upcoming_appointments,
                MAX(a.appointment_date) as last_visit_date
            FROM patients p
            LEFT JOIN appointments a ON a.patient_id = p.id
            GROUP BY p.id, p.first_name, p.last_name, p.gender, p.date_of_birth
            ORDER BY total_appointments DESC
        ''')

        # Gender distribution
        gender_dist = query('''
            SELECT 
                COALESCE(gender, 'Not Specified') as gender,
                COUNT(*) as count
            FROM patients
            GROUP BY gender
        ''')

        return ok({
            'summary': {
                'total_patients': len(patient_stats),
                'active_patients': sum(1 for p in patient_stats if p['total_appointments'] > 0)
            },
            'gender_distribution': gender_dist,
            'patients': patient_stats
        })
    except Exception as e:
        return fail(str(e), 500)


@app.route('/reports/revenue', methods=['GET'])
def get_revenue_report():
    """
    Get detailed revenue report with date range filter.
    Query params: start_date, end_date, group_by (day|week|month)
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        group_by = request.args.get('group_by', 'month')  # day, week, or month

        # Base query
        base_sql = '''
            SELECT 
                a.appointment_date,
                d.consultation_fee,
                s.name as speciality_name,
                d.first_name || ' ' || d.last_name as doctor_name
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            JOIN specialities s ON s.id = d.speciality_id
            WHERE a.status = 'completed'
        '''
        params = []

        if start_date:
            params.append(start_date)
            base_sql += f' AND a.appointment_date >= %s'
        if end_date:
            params.append(end_date)
            base_sql += f' AND a.appointment_date <= %s'

        base_sql += ' ORDER BY a.appointment_date'

        revenue_data = query(base_sql, tuple(params) if params else None)

        # Group by time period
        if group_by == 'day':
            date_format = 'YYYY-MM-DD'
            group_clause = "TO_CHAR(a.appointment_date, 'YYYY-MM-DD')"
        elif group_by == 'week':
            date_format = 'YYYY-WW'
            group_clause = "TO_CHAR(a.appointment_date, 'IYYY-IW')"
        else:  # month
            date_format = 'YYYY-MM'
            group_clause = "TO_CHAR(a.appointment_date, 'YYYY-MM')"

        grouped_sql = f'''
            SELECT 
                {group_clause} as period,
                COUNT(*) as appointment_count,
                COALESCE(SUM(d.consultation_fee), 0) as total_revenue
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            WHERE a.status = 'completed'
        '''

        if start_date:
            grouped_sql += f' AND a.appointment_date >= %s'
        if end_date:
            grouped_sql += f' AND a.appointment_date <= %s'

        grouped_sql += f' GROUP BY {group_clause} ORDER BY period'

        grouped_revenue = query(grouped_sql, tuple(params) if params else None)

        # Revenue by speciality
        speciality_revenue = query('''
            SELECT 
                s.name as speciality_name,
                COUNT(a.id) as appointment_count,
                COALESCE(SUM(d.consultation_fee), 0) as total_revenue
            FROM appointments a
            JOIN doctors d ON d.id = a.doctor_id
            JOIN specialities s ON s.id = d.speciality_id
            WHERE a.status = 'completed'
            GROUP BY s.name
            ORDER BY total_revenue DESC
        ''')

        # Total revenue
        total_revenue = sum(float(r['total_revenue']) for r in grouped_revenue)
        total_appointments = sum(r['appointment_count'] for r in grouped_revenue)

        return ok({
            'summary': {
                'total_revenue': total_revenue,
                'total_appointments': total_appointments,
                'average_per_appointment': round(total_revenue / total_appointments, 2) if total_appointments > 0 else 0
            },
            'revenue_by_period': grouped_revenue,
            'revenue_by_speciality': speciality_revenue,
            'details': revenue_data
        })
    except Exception as e:
        return fail(str(e), 500)


@app.route('/reports/export/<report_type>', methods=['GET'])
def export_report(report_type):
    """
    Export report in various formats.
    Query params: format (csv|json|excel), and report-specific filters
    """
    try:
        export_format = request.args.get('format', 'csv').lower()

        # Get data based on report type
        if report_type == 'appointments':
            # Get appointments report data
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')

            sql = '''
                SELECT 
                    a.id,
                    a.appointment_date,
                    a.start_time,
                    a.end_time,
                    a.status,
                    d.first_name || ' ' || d.last_name AS doctor_name,
                    p.first_name || ' ' || p.last_name AS patient_name,
                    s.name AS speciality_name,
                    d.consultation_fee
                FROM appointments a
                JOIN doctors d ON d.id = a.doctor_id
                JOIN patients p ON p.id = a.patient_id
                JOIN specialities s ON s.id = d.speciality_id
                WHERE 1=1
            '''
            params = []
            if start_date:
                params.append(start_date)
                sql += ' AND a.appointment_date >= %s'
            if end_date:
                params.append(end_date)
                sql += ' AND a.appointment_date <= %s'
            sql += ' ORDER BY a.appointment_date DESC'

            data = query(sql, tuple(params) if params else None)
            filename = f'appointments_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}'

        elif report_type == 'doctors':
            data = query('''
                SELECT 
                    d.id,
                    d.first_name || ' ' || d.last_name AS doctor_name,
                    s.name AS speciality_name,
                    d.email,
                    d.phone,
                    d.registration_number,
                    d.consultation_fee,
                    COUNT(a.id) as total_appointments,
                    COUNT(CASE WHEN a.status = 'completed' THEN 1 END) as completed_appointments,
                    COALESCE(SUM(CASE WHEN a.status = 'completed' THEN d.consultation_fee ELSE 0 END), 0) as total_revenue
                FROM doctors d
                JOIN specialities s ON s.id = d.speciality_id
                LEFT JOIN appointments a ON a.doctor_id = d.id
                WHERE d.is_active = TRUE
                GROUP BY d.id, d.first_name, d.last_name, s.name, d.email, d.phone, d.registration_number, d.consultation_fee
                ORDER BY total_revenue DESC
            ''')
            filename = f'doctors_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}'

        elif report_type == 'revenue':
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')

            sql = '''
                SELECT 
                    a.appointment_date,
                    d.first_name || ' ' || d.last_name as doctor_name,
                    s.name as speciality_name,
                    p.first_name || ' ' || p.last_name as patient_name,
                    d.consultation_fee as revenue,
                    a.status
                FROM appointments a
                JOIN doctors d ON d.id = a.doctor_id
                JOIN specialities s ON s.id = d.speciality_id
                JOIN patients p ON p.id = a.patient_id
                WHERE a.status = 'completed'
            '''
            params = []
            if start_date:
                params.append(start_date)
                sql += ' AND a.appointment_date >= %s'
            if end_date:
                params.append(end_date)
                sql += ' AND a.appointment_date <= %s'
            sql += ' ORDER BY a.appointment_date DESC'

            data = query(sql, tuple(params) if params else None)
            filename = f'revenue_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}'

        else:
            return fail('Invalid report type', 400)

        # Export based on format
        if export_format == 'json':
            return ok({'data': data, 'filename': filename})

        elif export_format == 'csv':
            if not data:
                return fail('No data to export', 404)

            # Create CSV in memory
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

            # Create response
            mem = io.BytesIO()
            mem.write(output.getvalue().encode('utf-8'))
            mem.seek(0)
            output.close()

            return send_file(
                mem,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'{filename}.csv'
            )

        elif export_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill

                if not data:
                    return fail('No data to export', 404)

                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = report_type.capitalize()

                # Add headers
                headers = list(data[0].keys())
                ws.append(headers)

                # Style headers
                for cell in ws[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Add data
                for row in data:
                    ws.append(list(row.values()))

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Save to memory
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'{filename}.xlsx'
                )
            except ImportError:
                return fail('Excel export requires openpyxl library. Install with: pip install openpyxl', 500)

        else:
            return fail('Invalid format. Use csv, json, or excel', 400)

    except Exception as e:
        return fail(str(e), 500)


if __name__ == '__main__':
    port = int(os.getenv('API_PORT', '5000'))
    app.run(debug=True, port=port)
